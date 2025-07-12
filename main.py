from flask import Flask, request, send_file, render_template
from utils import extract_text_from_pdf, compute_similarity
from sentence_transformers import SentenceTransformer
import csv
from io import BytesIO, StringIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import datetime
import pdfkit

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = "uploads"

model = SentenceTransformer("all-MiniLM-L6-v2")

# Path to wkhtmltopdf on Windows
path_wkhtmltopdf = r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
pdfkit_config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)


@app.route("/")
def landing_page():
    return render_template("landing.html")


@app.route("/screener", methods=["GET", "POST"])
def screener():
    results = []
    job_desc = ""
    if request.method == "POST":
        job_desc = request.form["job_desc"]
        files = request.files.getlist("resumes")

        for file in files:
            file_content = file.read()  # Read file into memory (bytes)
            text = extract_text_from_pdf(file_content)  # Use updated extractor

            score = compute_similarity(text, job_desc, model)
            score_percent = round(score * 100, 2)

            # Keyword logic
            summary = text[:250] + "..." if len(text) > 250 else text
            job_keywords = set(word.lower() for word in job_desc.split() if len(word) > 3)
            resume_words = set(word.lower() for word in text.split())
            matched = ", ".join(job_keywords & resume_words)
            missing = ", ".join(job_keywords - resume_words)
            job_role = "Software Engineer"

            results.append({
                "name": file.filename,
                "score": f"{score_percent}%",
                "text": text,
                "summary": summary,
                "matched": matched,
                "missing": missing,
                "job_role": job_role
            })

        results.sort(key=lambda r: float(r["score"].replace("%", "")), reverse=True)

    return render_template("index.html", results=results, job_desc=job_desc, job_role="Software Engineer")




@app.route("/download-pdf", methods=["POST"])
def download_pdf():
    job_desc = request.form["job_desc"]
    job_role = request.form.get("job_role", "Job")
    results_raw = request.form.getlist("results[]")
    texts = request.form.getlist("texts[]")
    summaries = request.form.getlist("summaries[]")
    keywords_matched = request.form.getlist("keywords_matched[]")
    keywords_missing = request.form.getlist("keywords_missing[]")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    result_data = []
    for i, row in enumerate(results_raw):
        name, score = row.split(" - ")
        try:
            score_val = float(score.strip().replace("%", ""))
        except ValueError:
            score_val = 0.0

        resume_name, _ = texts[i].split(":::", 1)

        result_data.append({
            "name": name.strip(),
            "score": round(score_val, 2),
            "matched": keywords_matched[i],
            "missing": keywords_missing[i],
            "summary": summaries[i]
        })

    html = render_template(
        "pdf_template.html",
        job_desc=job_desc,
        job_role=job_role,
        screening_date=timestamp,
        results=result_data,
        texts=texts,
        summaries=summaries,
        keywords_matched=keywords_matched,
        keywords_missing=keywords_missing
    )

    pdf = pdfkit.from_string(html, False, configuration=pdfkit_config)
    filename = f"{job_role.replace(' ', '_')}_results_{timestamp}.pdf"

    return send_file(BytesIO(pdf), download_name=filename, as_attachment=True, mimetype='application/pdf')


@app.route("/download-csv", methods=["POST"])
def download_csv():
    job_role = request.form.get("job_role", "Job")
    results = request.form.getlist("results[]")
    texts = request.form.getlist("texts[]")
    keywords_matched = request.form.getlist("keywords_matched[]")
    keywords_missing = request.form.getlist("keywords_missing[]")
    summaries = request.form.getlist("summaries[]")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    screening_date = datetime.now().strftime("%Y-%m-%d %H:%M")

    output = StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Resume", "Match Score (%)", "Top Keywords Matched", "Missing Keywords",
        "Summary", "Job Role", "Screening Date"
    ])

    for i, row in enumerate(results):
        name, score = row.split(" - ")
        resume_name, _ = texts[i].split(":::", 1)
        summary = summaries[i] if i < len(summaries) else ""
        short_summary = summary[:100] + "..." if len(summary) > 100 else summary

        writer.writerow([
            name.strip(),
            score.strip(),
            keywords_matched[i] if i < len(keywords_matched) else "",
            keywords_missing[i] if i < len(keywords_missing) else "",
            short_summary,
            job_role,
            screening_date
        ])

    byte_buffer = BytesIO(output.getvalue().encode("utf-8"))
    byte_buffer.seek(0)

    filename = f"{job_role.replace(' ', '_')}_results_{timestamp}.csv"
    return send_file(byte_buffer, mimetype='text/csv', download_name=filename, as_attachment=True)


@app.route("/download-xlsx", methods=["POST"])
def download_xlsx():
    job_role = request.form.get("job_role", "Job")
    results = request.form.getlist("results[]")
    keywords_matched = request.form.getlist("keywords_matched[]")
    keywords_missing = request.form.getlist("keywords_missing[]")
    summaries = request.form.getlist("summaries[]")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    screening_date = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "Resume", "Match Score (%)", "Top Keywords Matched", "Missing Keywords",
        "Summary", "Job Role", "Screening Date"
    ]
    ws.append(headers)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")

    for i, row in enumerate(results):
        try:
            name, score = row.split(" - ")
            score = score.strip().replace("%", "")
            numeric_score = (float(score))
        except ValueError:
            name = row
            numeric_score = 0

        summary = summaries[i] if i < len(summaries) else ""
        medium_summary = summary[:250] + "..." if len(summary) > 250 else summary
        matched = keywords_matched[i] if i < len(keywords_matched) else ""
        missing = keywords_missing[i] if i < len(keywords_missing) else ""

        ws.append([
            name.strip(),
            numeric_score,
            matched,
            missing,
            medium_summary,
            job_role,
            screening_date
        ])

        row_num = ws.max_row
        for col_index in range(1, 8):
            cell = ws.cell(row=row_num, column=col_index)
            cell.border = thin_border
            if row_num % 2 == 0:
                cell.fill = gray_fill
            if col_index in [3, 4, 5]:
                cell.alignment = Alignment(wrap_text=True)
            if col_index in [1, 2]:
                cell.alignment = Alignment(horizontal="center")
            if col_index == 2:
                cell.number_format = '0.00'

        if summary.strip():
            comment = Comment(text=medium_summary, author="AI Screener")
            ws.cell(row=row_num, column=2).comment = comment

        if missing.strip():
            comment = Comment(text=missing[:250] + "..." if len(missing) > 250 else missing, author="AI Screener")
            ws.cell(row=row_num, column=4).comment = comment

        if missing.strip():
            for col_index in range(1, 8):
                ws.cell(row=row_num, column=col_index).font = Font(bold=True)

    score_range = f"B2:B{ws.max_row}"
    ws.conditional_formatting.add(score_range, CellIsRule(operator='greaterThanOrEqual', formula=['85'],
                                                          fill=PatternFill(start_color="C6EFCE", fill_type="solid")))
    ws.conditional_formatting.add(score_range, CellIsRule(operator='between', formula=['60', '84'],
                                                          fill=PatternFill(start_color="FFF2CC", fill_type="solid")))
    ws.conditional_formatting.add(score_range, CellIsRule(operator='lessThan', formula=['60'],
                                                          fill=PatternFill(start_color="F8CBAD", fill_type="solid")))

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{job_role.replace(' ', '_')}_results_{timestamp}.xlsx"
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=filename, as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True)
